import os
import pandas as pd

from datetime import datetime
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter


class ReportGenerator:

    def __init__(
        self,
        profile,
        missing_report,
        duplicate_report,
        business_duplicate_report,
        validity_report,
        consistency_report,
        accuracy_report,
        dts_report,
        recommendations
    ):

        self.profile = profile
        self.missing_report = missing_report
        self.duplicate_report = duplicate_report
        self.business_duplicate_report = business_duplicate_report
        self.validity_report = validity_report
        self.consistency_report = consistency_report
        self.accuracy_report = accuracy_report
        self.dts_report = dts_report
        self.recommendations = recommendations

    def generate_report(
        self,
        path="data/output/DATATSTY_Enterprise_Report.xlsx"
    ):

        os.makedirs(
            os.path.dirname(path),
            exist_ok=True
        )

        with pd.ExcelWriter(
            path,
            engine="openpyxl"
        ) as writer:

            self.create_executive_summary().to_excel(
                writer,
                sheet_name="Executive_Summary",
                index=False
            )

            self.create_dataset_profile().to_excel(
                writer,
                sheet_name="Dataset_Profile",
                index=False
            )

            self.create_quality_findings().to_excel(
                writer,
                sheet_name="Quality_Findings",
                index=False
            )

            pd.DataFrame(
                self.recommendations
            ).to_excel(
                writer,
                sheet_name="Recommendations",
                index=False
            )

            workbook = writer.book

            for sheet in workbook.sheetnames:

                ws = workbook[sheet]

                self.apply_style(ws)

                self.auto_adjust_columns(ws)

                self.freeze_header(ws)

                if sheet == "Quality_Findings":

                    self.apply_severity_format(ws)

        print(
            "\nDATATSTY Enterprise Report Generated Successfully!"
        )

        return path

    def create_executive_summary(self):

        critical_issues = 0

        reports = [
            self.missing_report,
            self.validity_report,
            self.consistency_report,
            self.accuracy_report
        ]

        for report in reports:

            for item in report:

                if item["Severity"] == "Critical":

                    critical_issues += 1

        return pd.DataFrame([
            {
                "Audit Date":
                    datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),

                "Dataset":
                    self.profile["Dataset Name"],

                "Total Rows":
                    self.profile["Total Rows"],

                "Total Columns":
                    self.profile["Total Columns"],

                "DTS Score":
                    self.dts_report["DTS Score"],

                "AI Readiness":
                    self.dts_report["Status"],

                "Critical Issues":
                    critical_issues,

                "Completeness":
                    self.dts_report["Dimension Scores"]["Completeness"],

                "Uniqueness":
                    self.dts_report["Dimension Scores"]["Uniqueness"],

                "Validity":
                    self.dts_report["Dimension Scores"]["Validity"],

                "Consistency":
                    self.dts_report["Dimension Scores"]["Consistency"],

                "Accuracy":
                    self.dts_report["Dimension Scores"]["Accuracy"]
            }
        ])

    def create_dataset_profile(self):

        rows = []

        for key, value in self.profile.items():

            if key != "Column Details":

                rows.append({

                    "Metric": key,

                    "Value": value

                })

        return pd.DataFrame(rows)

    def create_quality_findings(self):

        findings = []

        for item in self.missing_report:

            findings.append({
                "Dimension": "Completeness",
                "Column": item["Column"],
                "Issue": "Missing Values",
                "Count": item["Missing Count"],
                "Percentage": item["Missing Percentage"],
                "Severity": item["Severity"]
            })

        findings.append({
            "Dimension": "Uniqueness",
            "Column": "Dataset",
            "Issue": "Duplicate Records",
            "Count": self.duplicate_report["Duplicate Records"],
            "Percentage": self.duplicate_report["Duplicate Percentage"],
            "Severity": self.duplicate_report["Severity"]
        })

        for item in self.business_duplicate_report:

            findings.append({
                "Dimension": "Business Key",
                "Column": item["Business Key"],
                "Issue": "Duplicate Business Key",
                "Count": item["Duplicate Records"],
                "Percentage": item["Duplicate Percentage"],
                "Severity": item["Severity"]
            })

        for item in self.validity_report:

            findings.append({
                "Dimension": "Validity",
                "Column": item["Column"],
                "Issue": "Invalid Values",
                "Count": item["Invalid Count"],
                "Percentage": item["Invalid Percentage"],
                "Severity": item["Severity"]
            })

        for item in self.consistency_report:

            findings.append({
                "Dimension": "Consistency",
                "Column": item["Column"],
                "Issue": item["Issue Type"],
                "Count": item["Affected Records"],
                "Percentage": item["Percentage"],
                "Severity": item["Severity"]
            })

        for item in self.accuracy_report:

            findings.append({
                "Dimension": "Accuracy",
                "Column": item["Column"],
                "Issue": "Outliers Detected",
                "Count": item["Outlier Count"],
                "Percentage": item["Outlier Percentage"],
                "Severity": item["Severity"]
            })

        return pd.DataFrame(findings)

    def apply_style(self, ws):

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for cell in ws[1]:

            cell.fill = header_fill
            cell.font = header_font

        for row in ws.iter_rows():

            for cell in row:

                cell.border = border

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

    def auto_adjust_columns(self, ws):

        for column in ws.columns:

            width = 0

            letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    width = max(
                        width,
                        len(str(cell.value))
                    )

                except Exception:
                    pass

            ws.column_dimensions[
                letter
            ].width = width + 2

    def freeze_header(self, ws):

        ws.freeze_panes = "A2"

    def apply_severity_format(self, ws):

        colors = {

            "Critical": "FF4C4C",

            "Medium": "FFD966",

            "Low": "C6E0B4",

            "None": "FFFFFF"
        }

        for row in ws.iter_rows():

            for cell in row:

                if cell.value in colors:

                    cell.fill = PatternFill(
                        start_color=colors[cell.value],
                        end_color=colors[cell.value],
                        fill_type="solid"
                    )